#!/usr/bin/env python3
"""Baseline_ENCE_ADIF.xlsx — baseline bottom-up de un ENCE con precios ADIF BPA 2026.
Hoja 'ENCE tipo' (BoQ unitario) + 'x5 + comparación' (proyecto y contraste)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/administrador/docker/LFC2/IX_WBS_Planificacion/Baseline_ENCE_ADIF.xlsx"

EURUSD = 1.09          # EUR -> USD
TRM    = 4400          # COP/USD
EURCOP = TRM * EURUSD  # 4796 COP/EUR
ING    = 0.30          # ingeniería + integración + pruebas + safety case (% sobre equipo)
usd = lambda eur: round(eur * EURUSD)
cop = lambda eur: round(eur * EURCOP)

NAVY, LIGHT = "0A192F", "EEF3FA"
f_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
f_sub   = Font(name="Calibri", size=9,  italic=True, color="FFFFFF")
f_hdr   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_cell  = Font(name="Calibri", size=10, color="1B2A41")
f_tag   = Font(name="Calibri", size=10, bold=True, color="1B2A41")
f_total = Font(name="Calibri", size=11, bold=True, color="0A3B33")
f_note  = Font(name="Calibri", size=8.5, italic=True, color="555555")
f_link  = Font(name="Calibri", size=9, color="1155CC", underline="single")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_hdr   = PatternFill("solid", fgColor="14304A")
fill_grp   = PatternFill("solid", fgColor="DDE7F2")
fill_sub   = PatternFill("solid", fgColor="DFF7F0")
fill_tot   = PatternFill("solid", fgColor="CFF3E6")
fill_pend  = PatternFill("solid", fgColor="FBE9D0")
fill_zebra = PatternFill("solid", fgColor=LIGHT)

thin = Side(style="thin", color="C9D3E0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")

EUR_FMT, EURU_FMT, COP_FMT = '#,##0 "€"', '#,##0.00 "€"', '#,##0 "M"'

def style_row(ws, r, n, font=f_cell, fill=None):
    for c in range(1, n + 1):
        cell = ws.cell(row=r, column=c); cell.font = font; cell.border = border
        if fill: cell.fill = fill

wb = Workbook()

# ════════════════ HOJA 1: ENCE TIPO ════════════════
ws = wb.active
ws.title = "ENCE tipo"
ws.sheet_view.showGridLines = False
for i, w in enumerate([20, 40, 14, 12, 7, 14, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:G1")
ws["A1"] = "BASELINE DE UN ENCE (tipo) — bottom-up con precios ADIF BPA"
ws["A1"].font = f_title; ws["A1"].fill = fill_title; ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:G2")
ws["A2"] = ("Enclavamiento electrónico (cantonamiento físico) · 1 de los 5 nodos (La Dorada–México, Pto Berrío–Grecia, "
            "Barrancabermeja, García Cadena, Zapatosa) · ADIF BPA v1.4.0 (19/02/2026, © ADIF 2025) · "
            "Cant. = ENCE promedio (Ardanuy ÷ 5) · 1 EUR = 4.796 COP")
ws["A2"].font = f_sub; ws["A2"].fill = fill_title

hdr_r = 4
for c, h in enumerate(["Categoría", "Concepto", "Código BPA", "€/ud", "Cant.", "€ total", "Estado"], 1):
    cell = ws.cell(row=hdr_r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if c in (4, 5, 6, 7) else left

# (cat, concepto, codigo, eur_ud, cant, estado)  estado: OK=ADIF real, PEND=estimado
rows = [
    ("Núcleo enclavamiento", "UCP — módulo central de proceso (grande)", "CAC020caa", 118006.02, 1, "OK"),
    ("Núcleo enclavamiento", "Armario de distribución de energía", "CAC010ca", 36806.72, 1, "OK"),
    ("Núcleo enclavamiento", "Bastidor de equipos de enclavamiento", "CAC030caa", 16399.36, 1, "OK"),
    ("Núcleo enclavamiento", "Cableado red local del enclavamiento", "CAC040ca", 5689.43, 1, "OK"),
    ("Núcleo enclavamiento", "Controlador de objetos 160E/128S", "CAC050ca", 33919.15, 1, "OK"),
    ("Núcleo enclavamiento", "Registrador jurídico", "CAC070", 23352.44, 1, "OK"),
    ("Núcleo enclavamiento", "Control de interfaz a CTC (serie)", "CAC080", 25505.19, 1, "OK"),
    ("Núcleo enclavamiento", "Firewall de enclavamiento", "CAC170", 6885.21, 1, "OK"),
    ("Aparatos de vía", "Motor eléctrico de aguja (con cerrojo)", "CFA010$", 10624.66, 8, "OK"),
    ("Aparatos de vía", "Comprobador de aguja", "CFA040$", 4963.85, 8, "OK"),
    ("Señalización", "Señal alta LED 3 focos", "CCA040$", 7934.23, 12, "OK"),
    ("Detección de tren", "Sensor de rueda contador de ejes", "CBB030$", 8286.12, 9, "OK"),
    ("Detección de tren", "Evaluador 8 contadores de ejes", "CBB040$", 13398.59, 2, "OK"),
    ("Armarios y cajas", "Armario de señalización (grande)", "CDB010$", 4151.33, 4, "OK"),
    ("Armarios y cajas", "Caja de terminales 50 bornas", "CDA010$", 1066.04, 10, "OK"),
    ("Arquitectura", "Caseta técnica prefabricada", "TMI010", 7180.94, 1, "OK"),
    ("Cableado", "Cable de señalización 7x4x0.9 (m)", "CEA030$", 11.67, 2000, "OK"),
    ("Cableado", "Cable F.O. 48 fibras canalización (m)", "TCJ010$", 5.48, 1000, "OK"),
    ("Energía", "SAI 20 kVA · 6h (autonomía DBCD ≥4h)", "COA220ckea", 32077.37, 1, "OK"),
]
r = hdr_r + 1
equipo = 0.0
for i, (cat, concepto, cod, eur_ud, cant, estado) in enumerate(rows):
    tot = eur_ud * cant; equipo += tot
    ws.cell(row=r, column=1, value=cat).font = f_tag
    ws.cell(row=r, column=2, value=concepto)
    ws.cell(row=r, column=3, value=cod)
    ws.cell(row=r, column=4, value=eur_ud).number_format = EURU_FMT
    ws.cell(row=r, column=5, value=cant)
    ws.cell(row=r, column=6, value=round(tot)).number_format = EUR_FMT
    ws.cell(row=r, column=7, value=("✅ ADIF" if estado == "OK" else "⚠ estimado"))
    style_row(ws, r, 7, fill=(fill_pend if estado == "PEND" else (fill_zebra if i % 2 else None)))
    ws.cell(row=r, column=1).alignment = left; ws.cell(row=r, column=2).alignment = left; ws.cell(row=r, column=3).alignment = left
    ws.cell(row=r, column=4).alignment = right; ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=6).alignment = right; ws.cell(row=r, column=7).alignment = center
    r += 1

def total_row(label, eur, fill, fmt_note=None):
    global r
    ws.cell(row=r, column=1, value=label)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value=round(eur)).number_format = EUR_FMT
    ws.cell(row=r, column=7, value=fmt_note or "")
    style_row(ws, r, 7, font=f_total, fill=fill)
    ws.cell(row=r, column=1).alignment = left; ws.cell(row=r, column=6).alignment = right
    ws.cell(row=r, column=7).font = f_note; ws.cell(row=r, column=7).alignment = center
    r += 1

ing = equipo * ING
total = equipo + ing
total_row("Subtotal EQUIPO (suministro + montaje)", equipo, fill_sub)
total_row(f"Ingeniería + integración + pruebas + safety case ({int(ING*100)}%)", ing, fill_sub)
total_row("TOTAL ENCE tipo", total, fill_tot, f"≈ ${usd(total):,} USD · ${cop(total)//1_000_000:,} M COP")
r += 1
for note in [
    "Cantidades = ENCE PROMEDIO (referencia Ardanuy ÷ 5 estaciones): 8 desvíos motorizados, 12 señales, 9 contadores de ejes. Ajustar por estación (La Dorada–México > Zapatosa).",
    "Precios ADIF = suministro + montaje, costes indirectos 6% incluidos. Referencia europea; aplicar factores LFC (trocha 914, logística, MO).",
    "Núcleo CAC 100% precios ADIF reales: UCP CAC020 + CAC010/030/040/050 + registrador CAC070 + interfaz CTC CAC080 + firewall CAC170. Catálogo completo (17 módulos) en Precios_Base_ADIF_BPA_2026.md.",
    "Módulos CAC por estación (no en el tipo): CAC060 interfaz a ENCE de otra tecnología (frontera FENOCO: García Cadena/Chiriguaná), CAC140 control de mando local (operación degradada), CAC090 rack adicional. Los módulos de mando CAC130-160 no se suman para no duplicar el controlador de objetos / equipo de campo.",
    "Detección por contador de ejes (Ardanuy). Alternativa: circuito de vía AF (CBC030$ €7.938/u) — solución mixta según TCO Detección.",
    "ENERGÍA: SAI COA220$ 6h/20 kVA (€32.077) cumple DBCD §10.5 (≥4h). COA200 (1h) y COA210 (2h) no alcanzan.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value=note).font = f_note; ws.cell(row=r, column=1).alignment = left
    r += 1

ENCE_EUR = total  # baseline por ENCE

# ════════════════ HOJA 2: x5 + COMPARACIÓN ════════════════
w2 = wb.create_sheet("x5 + comparación")
w2.sheet_view.showGridLines = False
for i, w in enumerate([34, 16, 14, 16, 30], 1):
    w2.column_dimensions[get_column_letter(i)].width = w

w2.merge_cells("A1:E1")
w2["A1"] = "ENCE — PROYECTO (5 nodos) Y CONTRASTE CON REFERENCIAS"
w2["A1"].font = f_title; w2["A1"].fill = fill_title; w2["A1"].alignment = Alignment(horizontal="left", vertical="center")
w2.row_dimensions[1].height = 26
w2.merge_cells("A2:E2")
w2["A2"] = "Baseline bottom-up ADIF vs Ardanuy y WBS. TRM 4.400 COP/USD · 1 EUR = 4.796 COP."
w2["A2"].font = f_sub; w2["A2"].fill = fill_title

r = 4
for c, h in enumerate(["Fuente / método", "$/ENCE (USD)", "$/ENCE (M COP)", "5 ENCE (M COP)", "Comentario"], 1):
    cell = w2.cell(row=r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if 2 <= c <= 4 else left
r += 1

ence_usd = usd(ENCE_EUR)
ence_cop_m = cop(ENCE_EUR) / 1_000_000
filas = [
    ("Bottom-up ADIF (este modelo)", ence_usd, ence_cop_m, ence_cop_m * 5,
     "Suma de partidas BPA reales + 30% ingeniería. Núcleo CAC 100% ADIF.", True),
    ("Ardanuy (consultor diseño)", 869377, 869377 * TRM / 1_000_000, 869377 * 5 * TRM / 1_000_000,
     "5 × $869.377 USD = $4,35M. Coincide con la Conceptual LFC.", False),
    ("WBS v3.0 (línea base antigua)", 181818, 800, 4000,
     "$800M COP/ENCE — subestima ~5× el equipo real. A corregir.", False),
    ("DT-COMS-2026-007 (LFC v2)", round(2000*1_000_000/TRM), 2000, 10000,
     "$2.000M COP/ENCE — entre el WBS viejo y el bottom-up real.", False),
]
for i, (fuente, u, cm, c5, com, hl) in enumerate(filas):
    w2.cell(row=r, column=1, value=fuente).font = (f_total if hl else f_cell)
    w2.cell(row=r, column=2, value=round(u)).number_format = '#,##0 "$"'
    w2.cell(row=r, column=3, value=round(cm)).number_format = COP_FMT
    w2.cell(row=r, column=4, value=round(c5)).number_format = COP_FMT
    w2.cell(row=r, column=5, value=com)
    style_row(w2, r, 5, fill=(fill_tot if hl else (fill_zebra if i % 2 else None)))
    w2.cell(row=r, column=1).alignment = left; w2.cell(row=r, column=5).alignment = left
    for cc in (2, 3, 4): w2.cell(row=r, column=cc).alignment = right
    r += 1

r += 1
for note in [
    f"LECTURA: el baseline bottom-up ADIF (~${ence_usd:,} USD/ENCE) CONFIRMA a Ardanuy ($869k) y expone que el WBS v3.0 ($800M COP) subestima el ENCE ~4–5×.",
    "Las cantidades por ENCE son promedio (Ardanuy ÷ 5). El desglose real por estación lo fija la ingeniería de detalle; La Dorada–México y Barrancabermeja serán mayores que Zapatosa.",
    "Para defensa ante Interventoría: cerrar los precios CAC010/030/040/050 en BPA y confirmar cantidades por estación (nº de desvíos, señales y contadores).",
]:
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    w2.cell(row=r, column=1, value=note).font = f_note; w2.cell(row=r, column=1).alignment = left
    r += 1

for sheet in (ws, w2):
    sheet.sheet_view.zoomScale = 110
wb.save(OUT)
print("OK ->", OUT)
print(f"ENCE tipo: EQUIPO €{equipo:,.0f} + ing 30% = TOTAL €{total:,.0f} = ${usd(total):,} USD = ${cop(total)/1e9:.2f}B COP")
print(f"5 ENCE: ${cop(total)*5/1e9:.2f}B COP = ${usd(total)*5/1e6:.2f}M USD")
