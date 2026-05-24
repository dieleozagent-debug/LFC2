#!/usr/bin/env python3
"""Costo_Apartaderos_PTC.xlsx — 1 apartadero · 10 apartaderos · Validación ADIF.
Comprobador desglosado (hardware/SIL-4/integración) + columna Referencia/Fuente."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/administrador/docker/LFC2/IX_WBS_Planificacion/Costo_Apartaderos_PTC.xlsx"

TRM = 4400      # COP/USD (Risk, estructuración)
AIU = 1.18      # factor AIU + IVA
usd = lambda cop_millones: round(cop_millones * 1_000_000 / TRM)

NAVY, LIGHT = "0A192F", "EEF3FA"
f_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
f_sub   = Font(name="Calibri", size=9,  italic=True, color="FFFFFF")
f_hdr   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_cell  = Font(name="Calibri", size=10, color="1B2A41")
f_tag   = Font(name="Calibri", size=10, bold=True, color="1B2A41")
f_ref   = Font(name="Calibri", size=8.5, color="555555")
f_total = Font(name="Calibri", size=11, bold=True, color="0A3B33")
f_note  = Font(name="Calibri", size=8.5, italic=True, color="555555")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_hdr   = PatternFill("solid", fgColor="14304A")
fill_cd    = PatternFill("solid", fgColor="DFF7F0")
fill_tot   = PatternFill("solid", fgColor="CFF3E6")
fill_grp   = PatternFill("solid", fgColor="F3EEDA")
fill_zebra = PatternFill("solid", fgColor=LIGHT)

thin = Side(style="thin", color="C9D3E0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")

COP_FMT, USD_FMT = '#,##0 "M"', '$#,##0'

def style_row(ws, r, ncols, font=f_cell, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = font; cell.border = border
        if fill: cell.fill = fill

wb = Workbook()

# ═══════════════════════ HOJA 1: 1 APARTADERO ═══════════════════════
ws1 = wb.active
ws1.title = "1 Apartadero"
ws1.sheet_view.showGridLines = False
for i, w in enumerate([18, 42, 13, 11, 46], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.merge_cells("A1:E1")
ws1["A1"] = "COSTO DE 1 APARTADERO  —  Sistema CTSC · Arquitectura PTC Virtual"
ws1["A1"].font = f_title; ws1["A1"].fill = fill_title; ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 26
ws1.merge_cells("A2:E2")
ws1["A2"] = "Corredor La Dorada–Chiriguaná · Contrato APP No. 001 de 2025 · SICC v14.7 · TRM 4.400 · AIU+IVA 1,18"
ws1["A2"].font = f_sub; ws1["A2"].fill = fill_title

hdr_r = 4
for c, h in enumerate(["Cubo", "Concepto", "COP (millones)", "USD", "Referencia / Fuente"], 1):
    cell = ws1.cell(row=hdr_r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if c in (3, 4) else left

# (cubo, concepto, COP_millones, referencia, es_subgrupo_comprobador)
rows1 = [
    ("Comunicaciones", "Terminal de datos TETRA ×2 + alta/programación en red", 35,
     "ADIF no tiene TETRA (usa GSM-R, TMG010 €2.727). Mercado Hytera/Motorola — RFQ pendiente", False),
    ("CTC/PTC", "Comprobador de aguja — hardware (suministro+montaje) ×2", 48,
     "ADIF CFA040$ €4.963,85/ud — bpa.adif.es", True),
    ("CTC/PTC", "Certificación SIL-4 + documentación safety case", 20,
     "FRA 49 CFR §236.1005 / EN 50126 — estimación", True),
    ("CTC/PTC", "Integración Back Office PTC + prueba funcional E2E (del punto)", 24,
     "AREMA C&S Manual §12 — estimación", True),
    ("Energía", "Solar autónoma <50 W + batería LiFePO4", 33,
     "Mercado Colombia solar industrial — RFQ pendiente", False),
    ("Servicios", "Instalación + prueba in-situ (SAT)", 31,
     "AREMA C&S §12 — rendimientos de campo (estimación)", False),
    ("Ingeniería", "Diseño de detalle + parametrización (prorrateado ÷10)", 90,
     "FRA §236.1005 — estimación % CAPEX", False),
]
r = hdr_r + 1
for i, (cubo, concepto, cop, ref, grp) in enumerate(rows1):
    ws1.cell(row=r, column=1, value=cubo).font = f_tag
    ws1.cell(row=r, column=2, value=("   • " + concepto) if grp else concepto).font = f_cell
    ws1.cell(row=r, column=3, value=cop).number_format = COP_FMT
    ws1.cell(row=r, column=4, value=usd(cop)).number_format = USD_FMT
    ws1.cell(row=r, column=5, value=ref).font = f_ref
    style_row(ws1, r, 5, fill=(fill_grp if grp else (fill_zebra if i % 2 else None)))
    ws1.cell(row=r, column=5).font = f_ref
    ws1.cell(row=r, column=1).alignment = left; ws1.cell(row=r, column=2).alignment = left
    ws1.cell(row=r, column=3).alignment = right; ws1.cell(row=r, column=4).alignment = right
    ws1.cell(row=r, column=5).alignment = left
    r += 1

cd1 = sum(c for _, _, c, _, _ in rows1)
ws1.cell(row=r, column=1, value="Costo directo (all-in)")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws1.cell(row=r, column=3, value=cd1).number_format = COP_FMT
ws1.cell(row=r, column=4, value=usd(cd1)).number_format = USD_FMT
ws1.cell(row=r, column=5, value="Comprobador (3 sublíneas) = subtotal CTC/PTC $92M")
style_row(ws1, r, 5, font=f_total, fill=fill_cd); ws1.cell(row=r, column=5).font = f_note
ws1.cell(row=r, column=3).alignment = right; ws1.cell(row=r, column=4).alignment = right; ws1.cell(row=r, column=5).alignment = left
cd_r = r; r += 1
ws1.cell(row=r, column=1, value="Con AIU + IVA")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws1.cell(row=r, column=3, value=round(cd1 * AIU, 1)).number_format = COP_FMT
ws1.cell(row=r, column=4, value=usd(cd1 * AIU)).number_format = USD_FMT
style_row(ws1, r, 5, font=f_total, fill=fill_tot)
ws1.cell(row=r, column=3).alignment = right; ws1.cell(row=r, column=4).alignment = right
r += 2
for note in [
    "Cifras de orden de magnitud — sujeto a RFQ formal (Frauscher / Siemens / Hytera).",
    "NO incluye: alargamiento de apartaderos (obra civil) ni infraestructura TETRA del corredor (ya presupuestada).",
    "CTC/PTC = comprobador hardware ($48M) + cert. SIL-4 ($20M) + integración ($24M) = $92M.",
    "Anclas de precio / códigos BPA: ver hoja 'Validación ADIF' (fuente bpa.adif.es · consulta 2026-05-23 · versión por confirmar).",
]:
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws1.cell(row=r, column=1, value=note).font = f_note; ws1.cell(row=r, column=1).alignment = left
    r += 1

# ═══════════════════════ HOJA 2: 10 APARTADEROS ═══════════════════════
ws2 = wb.create_sheet("10 Apartaderos")
ws2.sheet_view.showGridLines = False
for i, w in enumerate([18, 40, 7, 13, 11, 42], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:F1")
ws2["A1"] = "COSTO DE 10 APARTADEROS  —  Sistema CTSC · Arquitectura PTC Virtual"
ws2["A1"].font = f_title; ws2["A1"].fill = fill_title; ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 26
ws2.merge_cells("A2:F2")
ws2["A2"] = "Corredor La Dorada–Chiriguaná · APP No. 001 de 2025 · 10 apartaderos = 20 desvíos · TRM 4.400 · AIU+IVA 1,18"
ws2["A2"].font = f_sub; ws2["A2"].fill = fill_title

hdr_r = 4
for c, h in enumerate(["Cubo", "Concepto", "Cant.", "COP (millones)", "USD", "Referencia / Fuente"], 1):
    cell = ws2.cell(row=hdr_r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if c in (3, 4, 5) else left

rows2 = [
    ("Comunicaciones", "Terminal de datos TETRA + alta (infra TETRA ya presupuestada)", "20", 350,
     "ADIF usa GSM-R, no TETRA. Mercado Hytera/Motorola — RFQ", False),
    ("CTC/PTC", "Comprobador de aguja — hardware (suministro+montaje)", "20", 480,
     "ADIF CFA040$ €4.963,85/ud — bpa.adif.es", True),
    ("CTC/PTC", "Certificación SIL-4 + safety case", "10", 200,
     "FRA §236.1005 / EN 50126 — estimación", True),
    ("CTC/PTC", "Integración Back Office + prueba E2E", "10", 240,
     "AREMA C&S §12 — estimación", True),
    ("Energía", "Solar autónoma <50 W + LiFePO4", "10", 330,
     "Mercado Colombia solar industrial — RFQ", False),
    ("Servicios de campo", "Instalación + prueba in-situ (SAT)", "10", 310,
     "AREMA C&S §12 — rendimientos de campo", False),
    ("Ingeniería", "Diseño + config Back Office (global, una sola vez)", "1", 900,
     "FRA §236.1005 — estimación % CAPEX", False),
]
r = hdr_r + 1
for i, (cubo, concepto, cant, cop, ref, grp) in enumerate(rows2):
    ws2.cell(row=r, column=1, value=cubo).font = f_tag
    ws2.cell(row=r, column=2, value=("   • " + concepto) if grp else concepto).font = f_cell
    ws2.cell(row=r, column=3, value=cant)
    ws2.cell(row=r, column=4, value=cop).number_format = COP_FMT
    ws2.cell(row=r, column=5, value=usd(cop)).number_format = USD_FMT
    ws2.cell(row=r, column=6, value=ref).font = f_ref
    style_row(ws2, r, 6, fill=(fill_grp if grp else (fill_zebra if i % 2 else None)))
    ws2.cell(row=r, column=6).font = f_ref
    ws2.cell(row=r, column=1).alignment = left; ws2.cell(row=r, column=2).alignment = left
    ws2.cell(row=r, column=3).alignment = center; ws2.cell(row=r, column=4).alignment = right
    ws2.cell(row=r, column=5).alignment = right; ws2.cell(row=r, column=6).alignment = left
    r += 1

cd2 = sum(c for _, _, _, c, _, _ in rows2)
ws2.cell(row=r, column=1, value="Costo directo total")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws2.cell(row=r, column=4, value=cd2).number_format = COP_FMT
ws2.cell(row=r, column=5, value=usd(cd2)).number_format = USD_FMT
ws2.cell(row=r, column=6, value="Comprobador (3 sublíneas) = $920M")
style_row(ws2, r, 6, font=f_total, fill=fill_cd); ws2.cell(row=r, column=6).font = f_note
ws2.cell(row=r, column=4).alignment = right; ws2.cell(row=r, column=5).alignment = right; ws2.cell(row=r, column=6).alignment = left
cd_r = r; r += 1
ws2.cell(row=r, column=1, value="Con AIU + IVA")
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws2.cell(row=r, column=4, value=round(cd2 * AIU, 1)).number_format = COP_FMT
ws2.cell(row=r, column=5, value=usd(cd2 * AIU)).number_format = USD_FMT
style_row(ws2, r, 6, font=f_total, fill=fill_tot)
ws2.cell(row=r, column=4).alignment = right; ws2.cell(row=r, column=5).alignment = right
r += 2
for note in [
    "Cifras de orden de magnitud — sujeto a RFQ formal (Frauscher / Siemens / Hytera).",
    "NO incluye: alargamiento de apartaderos (obra civil) ni infraestructura TETRA del corredor (ya presupuestada en el proyecto).",
    "CTC/PTC = comprobador hardware ($480M) + cert. SIL-4 ($200M) + integración ($240M) = $920M.",
    "Infraestructura TETRA verificada en WBS v3.0: no hay partida de terminales para apartaderos → el terminal del switch es incremental (sin doble conteo).",
    "El costo unitario baja con el volumen: la ingeniería fija (~$900M) se reparte entre más apartaderos.",
    "Anclas de precio / códigos BPA: ver hoja 'Validación ADIF' (fuente bpa.adif.es · consulta 2026-05-23 · versión por confirmar).",
]:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws2.cell(row=r, column=1, value=note).font = f_note; ws2.cell(row=r, column=1).alignment = left
    r += 1

# ═══════════════════════ HOJA 3: VALIDACIÓN ADIF ═══════════════════════
ws3 = wb.create_sheet("Validación ADIF")
ws3.sheet_view.showGridLines = False
for i, w in enumerate([34, 17, 26, 11, 11, 34], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:F1")
ws3["A1"] = "VALIDACIÓN CONTRA ADIF — Banco de Precios (BPA)"
ws3["A1"].font = f_title; ws3["A1"].fill = fill_title; ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 26
ws3.merge_cells("A2:F2")
ws3["A2"] = ("Fuente: bpa.adif.es (consulta directa)  ·  Fecha de consulta: 2026-05-23  ·  "
             "Versión BPA: POR CONFIRMAR  ·  Conversión 4.796 COP/EUR (TRM 4.400 × 1,09)")
ws3["A2"].font = f_sub; ws3["A2"].fill = fill_title

EUR_FMT, COPM_FMT = '#,##0 "€"', '#,##0.0 "M"'
r = 4
ws3.cell(row=r, column=1, value="EXCLUIDOS — no son alcance del apartadero").font = f_total
r += 1
for c, h in enumerate(["Componente", "Código BPA", "Ruta BPA (capítulo)", "Precio €/ud", "≈ COP (M)", "Por qué se excluye"], 1):
    cell = ws3.cell(row=r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if 4 <= c <= 5 else left
r += 1
for i, (comp, cod, ruta, eur, cop, nota) in enumerate([
    ("Aparato de vía — desvío (suministro)", "VEA010aba", "V# › VE# › VEA# (Suministro)", 104566, 501.5, "Obra civil/vía. Apartaderos existentes (AT1). Montaje en VEC#."),
    ("Motor / accionamiento de aguja", "CFA010aaa", "C# › CF# (Aparatos vía) › CFA#", 10625, 51.0, "Autotalonable fuera de ENCE — sin motor."),
    ("Señal luminosa LED", "CCA040ebaad", "C# › CC# › CCA# (Señales laterales)", 7934, 38.1, "PTC virtual — sin semáforo lateral. Solo ENCE (WBS 1.5.101)."),
]):
    ws3.cell(row=r, column=1, value=comp); ws3.cell(row=r, column=2, value=cod); ws3.cell(row=r, column=3, value=ruta)
    ws3.cell(row=r, column=4, value=eur).number_format = EUR_FMT
    ws3.cell(row=r, column=5, value=cop).number_format = COPM_FMT
    ws3.cell(row=r, column=6, value=nota)
    style_row(ws3, r, 6, fill=(fill_zebra if i % 2 else None))
    for col in (1, 3, 6): ws3.cell(row=r, column=col).alignment = left
    ws3.cell(row=r, column=4).alignment = right; ws3.cell(row=r, column=5).alignment = right
    r += 1
r += 1
ws3.cell(row=r, column=1, value="INCLUIDOS — equipo CTSC incremental").font = f_total
r += 1
for c, h in enumerate(["Componente", "$M COP", "Ruta BPA / referencia", "", "", "Detalle"], 1):
    cell = ws3.cell(row=r, column=c, value=h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    cell.alignment = center if c == 2 else left
r += 1
for i, (comp, cop, ruta, detalle) in enumerate([
    ("Comprobador de aguja (×2) + integración CCO", 92, "ADIF CFA040caa · C# › CF# › CFA#",
     "€4.963,85/ud ≈ $23,8M/u (suministro+montaje). 2 ud ≈ $48M base; resto = SIL-4 vital + integración Back Office."),
    ("Terminal de datos TETRA (×2)", 35, "Mercado (ADIF no tiene TETRA, usa GSM-R)",
     "Ref. ADIF GSM-R TMG010a €2.727 (terminal mano). Mercado Hytera/Motorola ~$17,5M/u dato industrial."),
    ("Energía solar autónoma <50W + LiFePO4", 33, "Mercado solar local",
     "~$7.500 USD instalado."),
]):
    ws3.cell(row=r, column=1, value=comp); ws3.cell(row=r, column=2, value=cop).number_format = COP_FMT
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws3.cell(row=r, column=3, value=ruta); ws3.cell(row=r, column=6, value=detalle)
    style_row(ws3, r, 6, fill=(fill_zebra if i % 2 else None))
    ws3.cell(row=r, column=1).alignment = left; ws3.cell(row=r, column=2).alignment = right
    ws3.cell(row=r, column=3).alignment = left; ws3.cell(row=r, column=6).alignment = left
    r += 1
r += 1
for note in [
    "CORRECCIÓN: el código CBB010 no es el motor — en BPA es 'bastidor equipos de enclavamiento' (contadores de ejes). Motor = CFA010$.",
    "GSM-R vs TETRA: 'TETRA' da cero resultados en BPA — ADIF usa GSM-R. LFC usa TETRA por doctrina (BCD) -> terminales por mercado.",
    "Abatibles: la distinción abatible/no abatible existe en CCA040$ (parámetro BPA), pero solo aplica a señales de ENCE.",
    "El comprobador es la línea más sensible: rango ADIF base ~$23,8M/u <-> versión SIL-4 vital. Confirmar vía RFQ.",
    "PENDIENTE: capturar la VERSIÓN/año exacta del BPA y la URL directa por ítem en la próxima consulta (no se registraron).",
]:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws3.cell(row=r, column=1, value=note).font = f_note; ws3.cell(row=r, column=1).alignment = left
    r += 1

for ws in (ws1, ws2, ws3):
    ws.sheet_view.zoomScale = 110
wb.save(OUT)
print("OK ->", OUT)
