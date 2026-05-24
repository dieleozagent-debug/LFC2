#!/usr/bin/env python3
"""Valida Presupuesto_SCC_LFC2 contra precios ADIF BPA v1.4.0 → hoja '8.Validación ADIF'."""
import glob, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = sorted(glob.glob("X_ENTREGABLES_CONSOLIDADOS/Presupuesto_SCC_LFC2_*.xlsx"))[-1]
OUT = "X_ENTREGABLES_CONSOLIDADOS/Presupuesto_SCC_LFC2_validado.xlsx"
EURCOP = 4796.0
BP = "https://bpa.adif.es/bp1/#/concept/"

# item -> dict(code, eur, url, note, unit('ud'|'km'), combine(item), scope(bool))
REF = {}
for it in ("1.3.100", "1.3.101", "1.3.102", "1.3.103", "1.3.104"):
    REF[it] = dict(code="núcleo CAC + módulos objeto", eur=414305, url=BP+"152945939",
                   note='ENCE completo (núcleo €249.783 + módulos €164.522). Nota "CAC020 €356.780" INCORRECTA → real €118.006.')
REF["1.4.100"] = dict(code="VEA010+CFA010+CFA040+CFA030", eur=122397, url=BP+"156706377",
                      note="Suministro €104.566 + motor + comprobador + cerrojo. Montaje y trocha 914mm justifican el delta.")
REF["1.5.100"] = dict(code="CFA040caa (+ switch)", eur=4963.85, url=BP+"151681024", scope=True,
                      note="BPA cubre solo el comprobador ($23,8M). El VU incluye el switch autotalonable completo + telemetría → alcance distinto, no comparable directo.")
REF["1.5.101"] = dict(code="CCA040ebaad", eur=7934.23, url=BP+"153988522",
                      note="Señal alta LED 3 focos €7.934. Delta por poste + montaje + factor Colombia.")
REF["2.3.100"] = dict(code="TCJ010bccca", eur=5.48*1000, url=BP+"158651112", combine="2.3.101", unit="km",
                      note="BPA €5,48/m incl. tendido (=€5.480/km). Se compara la SUMA supply(2.3.100)+MO(2.3.101) vs BPA.")
# notas de analogía (sin delta) para ítems sin partida directa
ANALOG = {
    "2.1.106": "Analogía ADIF caseta TMI010 €7.181 ($34,4M); caseta TETRA difiere → RFQ.",
    "2.1.116": "ADIF SAI máx 6h (COA220 €32.077); autonomía 24h TETRA fuera de BPA → RFQ.",
    "2.3.103": "Cajas empalme FO; sin partida BPA directa → RFQ.",
}

wb = openpyxl.load_workbook(SRC)            # con fórmulas/estilos
wbv = openpyxl.load_workbook(SRC, data_only=True)  # valores calculados
ws = wb["3.Detalle Ítems"]; wsv = wbv["3.Detalle Ítems"]

# estilos
NAVY = "0A192F"
f_t = Font(size=13, bold=True, color="FFFFFF"); f_h = Font(size=10, bold=True, color="FFFFFF")
f_c = Font(size=9, color="1B2A41"); f_n = Font(size=8.5, color="555555")
fill_t = PatternFill("solid", fgColor=NAVY); fill_h = PatternFill("solid", fgColor="14304A")
GREEN = PatternFill("solid", fgColor="C6EFCE"); YEL = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE"); GREY = PatternFill("solid", fgColor="E2E2E2")
thin = Side(style="thin", color="C9D3E0"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
L = Alignment("left", "center", wrap_text=True); R = Alignment("right", "center"); C = Alignment("center", "center")

if "8.Validación ADIF" in wb.sheetnames:
    del wb["8.Validación ADIF"]
v = wb.create_sheet("8.Validación ADIF")
v.sheet_view.showGridLines = False
for i, w in enumerate([10, 40, 16, 12, 22, 12, 9, 12, 40, 30], 1):
    v.column_dimensions[chr(64+i)].width = w
v.merge_cells("A1:J1")
v["A1"] = "VALIDACIÓN CONTRA ADIF BPA — v1.4.0 · válida 19/02/2026 · consulta 23/05/2026 · TRM 4.400 · EUR/COP 4.796"
v["A1"].font = f_t; v["A1"].fill = fill_t; v["A1"].alignment = Alignment("left", "center")
v.row_dimensions[1].height = 24
hdr = ["Ítem", "Descripción", "VU Excel (COP)", "VU Excel (€)", "Ref BPA (código)",
       "Ref BPA (€)", "Delta %", "Estado", "Nota", "URL BPA"]
for c, h in enumerate(hdr, 1):
    cell = v.cell(3, c, h); cell.font = f_h; cell.fill = fill_h; cell.border = border
    cell.alignment = C if c in (3, 4, 6, 7, 8) else L

def status(delta):
    if -15 <= delta <= 40: return "✅ OK", GREEN
    if 40 < delta <= 100 or -30 <= delta < -15: return "⚠ revisar", YEL
    return "🔴 fuera de rango", RED

r = 4
n_ok = n_warn = n_red = n_sinref = 0
for row in range(2, wsv.max_row + 1):
    item = wsv.cell(row, 3).value
    desc = wsv.cell(row, 4).value
    vu = wsv.cell(row, 8).value
    if not item or vu in (None, "", 0):
        continue
    try:
        vu = float(vu)
    except (TypeError, ValueError):
        continue
    item = str(item).strip()
    vu_eur = vu / EURCOP
    ref = REF.get(item)
    nota = ""; url = ""; code = ""; ref_eur = ""; delta_txt = ""; est = ""; fill = GREY
    if ref:
        code = ref["code"]; ref_eur = ref["eur"]; url = ref["url"]; nota = ref["note"]
        cmp_vu = vu
        if ref.get("combine"):
            # suma con el ítem combinado (mismo unit)
            for rr in range(2, wsv.max_row + 1):
                if str(wsv.cell(rr, 3).value).strip() == ref["combine"]:
                    cmp_vu = vu + float(wsv.cell(rr, 8).value or 0); break
        cmp_eur = cmp_vu / EURCOP
        if ref.get("scope"):
            est = "⚠ alcance distinto"; fill = YEL; delta_txt = "n/a"; n_warn += 1
        else:
            delta = (cmp_eur - ref_eur) / ref_eur * 100
            delta_txt = f"{delta:+.0f}%"
            est, fill = status(delta)
            if fill is GREEN: n_ok += 1
            elif fill is YEL: n_warn += 1
            else: n_red += 1
    else:
        code = "—"; ref_eur = ""; est = "⬜ SIN REF"; fill = GREY
        nota = ANALOG.get(item, "Sin partida BPA directa → RFQ pendiente.")
        n_sinref += 1

    v.cell(r, 1, item).font = f_c
    v.cell(r, 2, str(desc)[:80]).font = f_c
    v.cell(r, 3, round(vu)).number_format = '#,##0'
    v.cell(r, 4, round(vu_eur)).number_format = '#,##0 "€"'
    v.cell(r, 5, code).font = f_c
    v.cell(r, 6, ref_eur if ref_eur == "" else round(ref_eur)).number_format = '#,##0 "€"'
    v.cell(r, 7, delta_txt)
    v.cell(r, 8, est)
    v.cell(r, 9, nota).font = f_n
    uc = v.cell(r, 10, url)
    if url:
        uc.hyperlink = url; uc.font = Font(size=8, color="1155CC", underline="single")
    for c in range(1, 11):
        cell = v.cell(r, c); cell.border = border
        if c in (3, 4, 6, 7): cell.alignment = R
        elif c == 8: cell.alignment = C; cell.fill = fill; cell.font = Font(size=9, bold=True)
        else: cell.alignment = L
    r += 1

# resumen al pie
r += 1
v.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
v.cell(r, 1, f"Resumen: ✅ {n_ok} OK · ⚠ {n_warn} revisar · 🔴 {n_red} fuera de rango · ⬜ {n_sinref} sin ref (RFQ). "
             "Solo ítems con partida ADIF mapeada se contrastan; el resto requiere RFQ. No se modificó ningún VU.").font = f_n

# --- corrección de notas de descripción (solo texto) en 3.Detalle Ítems ---
fixes = 0
for row in range(2, ws.max_row + 1):
    item = str(ws.cell(row, 3).value or "").strip()
    d = ws.cell(row, 4)
    if item in ("1.3.100", "1.3.101", "1.3.102", "1.3.103", "1.3.104") and d.value and "356" in str(d.value):
        d.value = str(d.value).split("(DT-COMS")[0].strip() + \
            " (BPA v1.4.0: núcleo CAC €249.783 + módulos ≈ €414K = $1.987M COP; UCP real CAC020caa €118.006, no €356.780)"
        fixes += 1
    if item == "1.4.100" and d.value and "114" in str(d.value):
        d.value = str(d.value).split("(DT-COMS")[0].strip() + \
            " (BPA v1.4.0: VEA010aba €104.566 solo suministro; +montaje +trocha 914mm justifican $640M COP)"
        fixes += 1

wb.save(OUT)
print(f"Fuente: {os.path.basename(SRC)}")
print(f"Guardado: {OUT}")
print(f"Ítems validados: OK={n_ok} · revisar={n_warn} · fuera={n_red} · sin_ref={n_sinref}")
print(f"Notas de descripción corregidas: {fixes}")
